from email import header
import sqlite3
import logging.handlers
import os
import csv
from os import path
import datetime
import json
from pathlib import Path
import typer
from openpyxl import Workbook

def to_csv(rows: list, header: list, output_file: str):
    try:
        with open(output_file, 'w') as f:
            csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator='\n').writerows([header])
            csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator='\n').writerows(rows)

    except Exception as sub_ex:
        log_dog.error(sub_ex)
        raise


def to_excel(rows: list, header: list, output_file: str):
    try:
        wb = Workbook()
        sheet = wb.active

        for column in header:
            for col_ctr in range(1, len(header) + 1):
                sheet.cell(row=1, column=col_ctr).value = column


        row_ctr = 2
        for row in rows:
            col_ctr = 1
            for col in row:
                if type(col) == 'str':
                    sheet.cell(row=row_ctr, column=col_ctr).value = col.strip()
                else:
                    sheet.cell(row=row_ctr, column=col_ctr).value = col     
                col_ctr += 1
            row_ctr += 1

        wb.save(output_file)
        wb.close()

    except Exception as sub_ex:
        log_dog.error(sub_ex)
        raise

def get_connection(db: str) -> sqlite3.Connection:
    try:
        return sqlite3.connect(db)

    except Exception as sub_ex:
        log_dog.error(sub_ex)
        raise


def get_cursor(connection: sqlite3.Connection) -> sqlite3.Cursor:
    try:
        return connection.cursor()

    except Exception as sub_ex:
        log_dog.error(sub_ex)
        raise


def to_db(rows: list, db: str):
    try:
        if not os.path.exists(db):
            raise ValueError(f"{db} does not exist")

        dbcon = get_connection(db)
        cursor = get_cursor(dbcon)
        
        print('processing db')

        # clear the table which also checks for tables existance
        sql = "delete from files"
        cursor.execute(sql)
        dbcon.commit()

        sql = "insert into files (folder_file_name, name, suffix, size, modified, parts) values (?,?,?,?,?,?)"
        
        for row in rows:
            cursor.execute(sql, row)

        dbcon.commit()

    except Exception as sub_ex:
        log_dog.error(sub_ex)
    
        # if ftp_errors:
        #        raise ValueError(ftp_errors)


def set_up_error_log(log_file: str) -> object:    
    err_log = logging.getLogger(__name__)
    err_log.setLevel(logging.ERROR)

    bundle_dir = path.abspath(path.dirname(__file__))
    path_to_log = path.join(bundle_dir, log_file)
    
    print(path_to_log)

    h1 = logging.FileHandler(log_file)
    f = logging.Formatter("%(levelname)s %(asctime)s %(funcName)s %(lineno)d %(message)s")
    h1.setFormatter(f)
    h1.setLevel(logging.ERROR)
    err_log.addHandler(h1)

    return err_log


def get_folder_listing(folder : str) -> tuple:
    header = ['folder_file_name', 'name', 'suffix', 'size', 'modified', 'parts'] 
    rows = []
    for base_folder, _, f_iles in os.walk(folder):
        for file in f_iles:
            file_obj = Path(os.path.join(base_folder, file))
            stem = file_obj.stem
            suffix = file_obj.suffix
            parts = json.dumps(file_obj.parts)
            stat = file_obj.stat()
            size = stat.st_size 
            modified = datetime.datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
            folder_file = os.path.join(base_folder, file)
            rows.append([folder_file, stem, suffix, size, modified, parts])

    return rows, header

def output_callback(value: str):
    if value not in  ['csv', 'xl', 'xls', 'xlsx', 'db']:
        raise typer.BadParameter("Only csv, xl, xls, xlsx, db are allowed ouptut types")
    return value

def file_callback(ctx: typer.Context, param: typer.CallbackParam, value: str):
    # print(dir(param))
    # print(param.name)
    # print(ctx)
    # print(dir(ctx))
    if 'output' not in ctx.params:
        raise typer.BadParameter(f"output type  must be before file name")

    # print(ctx.params['output'])
    if ctx.params['output'] == 'db':
        if not os.path.exists(value):
            raise typer.BadParameter(f"{value} - db file must exist")

    return value

log_dog =  set_up_error_log('folderlist_error.log')


def main(
    folder: str = typer.Option(..., "--folder", "-f", help="Path to folder"),
    output: str = typer.Option("csv", "--output", "-ot", help="Output Type csv xls db", callback=output_callback),
    file: str = typer.Option(..., "--outfile", "-of", help="File Name and Path", callback=file_callback)
):

    try:
        rows, header = get_folder_listing(folder)
              
        if output == 'csv':
            to_csv(rows, header, file)
        elif output in ('xl', 'xls', 'xlsx'):
            to_excel(rows, header, file)
        else:
            to_db(rows, file)    


    except Exception as sub_ex:
        log_dog.exception('Main Process Exceptions:')



if __name__ == "__main__":
    typer.run(main)